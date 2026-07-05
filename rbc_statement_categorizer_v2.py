"""
RBC Business Statement Categorizer — v2
-----------------------------------------
Uses word-position coordinates to extract transactions
from RBC PDFs that don't have detectable table structures.

Requirements:
    pip install pdfplumber openpyxl

How to use:
    Double-click this file, drag your unlocked PDF folder in, press Enter.
"""

import sys
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMIZE YOUR CATEGORIES HERE
# ─────────────────────────────────────────────────────────────────────────────
CATEGORY_RULES = [
    ("Income",              ["MISC PAYMENT", "MOBILE CHEQUE DEPOSIT", "BR TO BR",
                              "E-TRANSFER RECEIVED", "PAYPAL"]),
    ("Telecom & Internet",  ["BELL", "ROGERS", "TELUS", "FIDO", "KOODO", "SHAW", "COGECO"]),
    ("Bank Fees",           ["MONTHLY FEE", "TRANSACTION FEE", "INTERAC",
                              "ELECTRONIC TRANSACTION", "REGULAR TRANSACTION",
                              "SERVICE CHARGE", "NSF", "PAY-FILE FEES",
                              "ACCOUNT FEE", "INSURANCE LOAN"]),
    ("e-Transfers Out",     ["E-TRANSFER SENT"]),
    ("Cheques Issued",      ["CHEQUE"]),
    ("Online Payments",     ["ONLINE BANKING PAYMENT"]),
    ("Government & Tax",    ["CRA", "CANADA REVENUE", "HST", "GST", "WSIB"]),
    ("Insurance",           ["INTACT", "AVIVA", "ECONOMICAL", "DESJARDINS",
                              "CO-OPERATORS", "FORESTERS", "DEFINITY",
                              "FIRST INSURANCE"]),
    ("Utilities",           ["HYDRO", "ENBRIDGE", "UNION GAS", "WATER"]),
    ("Office & Supplies",   ["STAPLES", "OFFICE DEPOT", "AMAZON", "MICROSOFT", "GOOGLE"]),
    ("Fuel & Auto",         ["PETRO", "SHELL", "ESSO", "CANADIAN TIRE", "ULTRAMAR"]),
    ("Meals & Travel",      ["TIM HORTONS", "STARBUCKS", "MCDONALD", "UBER", "HOTEL"]),
    ("Credit Card Payment", ["MASTERCD", "VISA", "AMEX", "MASTERCARD"]),
]
# ─────────────────────────────────────────────────────────────────────────────

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

SKIP_LINES = {
    "openingbalance", "closingbalance", "accountfees",
    "date", "description", "accountactivitydetails",
    "accountactivitydetails-continued", "1of", "2of", "3of",
}

# ── Check dependencies ────────────────────────────────────────────────────────
missing = []
try:
    import pdfplumber
except ImportError:
    missing.append("pdfplumber")

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    missing.append("openpyxl")

if missing:
    print("\n  Missing libraries. Open Command Prompt and run:")
    for lib in missing:
        print(f"      pip install {lib}")
    input("\nPress Enter to close...")
    sys.exit(1)


# ── Helpers ───────────────────────────────────────────────────────────────────

def categorize(description: str) -> str:
    upper = description.upper()
    for category, keywords in CATEGORY_RULES:
        if any(kw.upper() in upper for kw in keywords):
            return category
    return "Other"


def clean_amount(text: str) -> float:
    cleaned = re.sub(r"[,$\s]", "", text)
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_statement_period(text: str):
    """Handle both 'December16,2021toJanuary14,2022' and 'December 16, 2021 to January 14, 2022'."""
    # Add spaces around 'to' for easier parsing
    text = re.sub(r'([a-z])to([A-Z])', r'\1 to \2', text)

    pattern = r'([A-Za-z]+)\s*(\d{1,2}),?\s*(\d{4})\s+to\s+([A-Za-z]+)\s*(\d{1,2}),?\s*(\d{4})'
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        m1, d1, y1, m2, d2, y2 = match.groups()
        period = f"{m1} {d1}, {y1} to {m2} {d2}, {y2}"
        return period, int(y2)

    year_match = re.search(r'20(\d{2})', text)
    year = int("20" + year_match.group(1)) if year_match else datetime.now().year
    return "Unknown Period", year


def parse_rbc_date(date_str: str, year: int, prev_month: int) -> tuple[str, int]:
    """
    Parse dates like '20Dec', '04Jan', '20 Dec'.
    Returns (YYYY-MM-DD string, month_number)
    Handles year rollover (e.g. Dec statement with Jan transactions = next year).
    """
    date_str = date_str.strip()
    match = re.match(r'^(\d{1,2})\s*([A-Za-z]{3})$', date_str)
    if not match:
        return "", prev_month

    day = int(match.group(1))
    month_str = match.group(2).lower()
    month = MONTH_MAP.get(month_str, 0)

    if month == 0:
        return "", prev_month

    # Handle year rollover: if previous month was Dec and current is Jan
    tx_year = year
    if prev_month == 12 and month == 1:
        tx_year = year + 1
    elif prev_month == 1 and month == 12:
        tx_year = year - 1

    try:
        dt = datetime(tx_year, month, day)
        return dt.strftime("%Y-%m-%d"), month
    except ValueError:
        return "", prev_month


def extract_words_as_rows(page):
    """
    Extract words with x,y coordinates and group into rows by y-position.
    Returns list of rows, each row is list of (text, x0, x1) tuples sorted by x.
    """
    words = page.extract_words(x_tolerance=4, y_tolerance=4, keep_blank_chars=False)
    if not words:
        return []

    # Group by y (top position)
    row_dict = defaultdict(list)
    for w in words:
        y_bucket = round(w['top'] / 4) * 4
        row_dict[y_bucket].append((w['text'], w['x0'], w['x1']))

    # Sort rows by y, sort words within each row by x
    rows = []
    for y in sorted(row_dict.keys()):
        row_words = sorted(row_dict[y], key=lambda w: w[1])
        rows.append(row_words)

    return rows


def find_column_x(page):
    """
    Locate the x-position of Debits, Credits, Balance columns
    by finding the header row.
    Falls back to page-width percentages if header not found.
    """
    pw = page.width
    defaults = {
        'debit':   pw * 0.52,
        'credit':  pw * 0.68,
        'balance': pw * 0.82,
    }

    words = page.extract_words(x_tolerance=4, y_tolerance=4)
    header_words = {}
    for w in words:
        t = w['text'].lower()
        if t in ('cheques', 'debits') and 'debit' not in header_words:
            header_words['debit'] = w['x0']
        elif t in ('deposits', 'credits') and 'credit' not in header_words:
            header_words['credit'] = w['x0']
        elif t == 'balance' and 'balance' not in header_words:
            header_words['balance'] = w['x0']

    return {
        'debit':   header_words.get('debit',   defaults['debit']),
        'credit':  header_words.get('credit',  defaults['credit']),
        'balance': header_words.get('balance', defaults['balance']),
    }


def is_amount(text: str) -> bool:
    return bool(re.match(r'^\$?\d{1,3}(?:,\d{3})*\.\d{2}$', text))


def row_to_transaction(row_words, col_x, last_date, statement_year, prev_month):
    """
    Given a list of (text, x0, x1) for one row, extract a transaction dict.
    Returns (transaction_dict_or_None, updated_last_date, updated_prev_month)
    """
    if not row_words:
        return None, last_date, prev_month

    # Separate words into columns by x-position
    date_parts   = []
    desc_parts   = []
    debit_parts  = []
    credit_parts = []
    bal_parts    = []

    debit_x  = col_x['debit']
    credit_x = col_x['credit']
    bal_x    = col_x['balance']
    desc_x   = debit_x  # description ends where debit column starts

    for text, x0, x1 in row_words:
        mid_x = (x0 + x1) / 2
        if mid_x >= bal_x:
            bal_parts.append(text)
        elif mid_x >= credit_x:
            credit_parts.append(text)
        elif mid_x >= debit_x:
            debit_parts.append(text)
        elif mid_x >= 60:   # description zone (skip far-left page numbers)
            desc_parts.append(text)
        else:
            date_parts.append(text)

    date_str   = " ".join(date_parts).strip()
    desc_str   = " ".join(desc_parts).strip()
    debit_str  = " ".join(debit_parts).strip()
    credit_str = " ".join(credit_parts).strip()

    # Skip non-transaction rows
    combined_lower = (date_str + desc_str).lower().replace(" ", "")
    if any(combined_lower.startswith(skip) for skip in SKIP_LINES):
        return None, last_date, prev_month
    if not desc_str:
        return None, last_date, prev_month

    # Try to parse date
    new_date = last_date
    new_month = prev_month

    # Date might be in date_parts or start of desc_parts
    candidate_date = date_str
    if not candidate_date:
        # Maybe date is first word of description
        first_word = desc_parts[0] if desc_parts else ""
        if re.match(r'^\d{1,2}[A-Za-z]{3}$', first_word):
            candidate_date = first_word
            desc_str = " ".join(desc_parts[1:]).strip()

    if candidate_date:
        parsed, parsed_month = parse_rbc_date(candidate_date, statement_year, prev_month)
        if parsed:
            new_date = parsed
            new_month = parsed_month

    if not new_date:
        return None, last_date, prev_month
    if not desc_str:
        return None, new_date, new_month

    # Parse amounts
    debit  = clean_amount(debit_str)  if is_amount(debit_str)  else 0.0
    credit = clean_amount(credit_str) if is_amount(credit_str) else 0.0
    amount = credit - debit

    # Skip rows with no monetary value
    if debit == 0.0 and credit == 0.0:
        # Still keep bank fee description lines
        if not any(w in desc_str.upper() for w in ["FEE", "CHARGE", "LOAN"]):
            return None, new_date, new_month

    tx = {
        "Date":        new_date,
        "Description": desc_str,
        "Debit":       debit  if debit  > 0 else None,
        "Credit":      credit if credit > 0 else None,
        "Amount":      amount,
        "Category":    categorize(desc_str),
    }

    return tx, new_date, new_month


def extract_transactions_from_pdf(pdf_path: Path):
    """Main extraction function. Returns (period_str, [transactions])."""
    transactions = []
    period_str = "Unknown Period"
    statement_year = datetime.now().year

    with pdfplumber.open(str(pdf_path)) as pdf:
        # Get full text to parse statement period
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

        period_str, statement_year = parse_statement_period(full_text)

        last_date  = ""
        prev_month = 0

        for page in pdf.pages:
            col_x = find_column_x(page)
            rows  = extract_words_as_rows(page)

            for row_words in rows:
                tx, last_date, prev_month = row_to_transaction(
                    row_words, col_x, last_date, statement_year, prev_month
                )
                if tx:
                    tx["Period"] = period_str
                    tx["File"]   = pdf_path.name
                    transactions.append(tx)

    return period_str, transactions


# ── Excel Output ──────────────────────────────────────────────────────────────

def solid(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def cfont(size=10, bold=False, color="1A1A2E"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def border():
    s = Side(style="thin", color="DEE2E6")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center")

def vcenter():
    return Alignment(vertical="center")

NAVY  = "1A1A2E"
BLUE  = "0F3460"
GREEN_BG = "EDF7F0"
RED_BG   = "FDF0F0"
LIGHT    = "F8F9FA"
WHITE    = "FFFFFF"
GREEN_FG = "1B6B3A"
RED_FG   = "C0392B"


def write_transactions_sheet(ws, transactions):
    ws.title = "All Transactions"
    headers = ["Date", "Description", "Category", "Debit ($)", "Credit ($)", "Net ($)", "Period"]
    widths  = [14, 46, 24, 14, 14, 14, 38]

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "RBC Business Account — Full Transaction History"
    c.font  = hfont(size=13)
    c.fill  = solid(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # Headers
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font      = hfont(size=10)
        cell.fill      = solid(BLUE)
        cell.alignment = center()
        cell.border    = border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    sorted_tx = sorted(transactions, key=lambda x: x.get("Date", "9999"))

    for r, tx in enumerate(sorted_tx, start=3):
        amt = tx.get("Amount", 0)
        bg  = solid(GREEN_BG) if amt > 0 else (solid(RED_BG) if amt < 0 else solid(LIGHT))

        vals = [
            tx.get("Date", ""),
            tx.get("Description", ""),
            tx.get("Category", ""),
            tx.get("Debit"),
            tx.get("Credit"),
            amt if amt != 0 else None,
            tx.get("Period", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font   = cfont()
            cell.fill   = bg
            cell.border = border()
            cell.alignment = vcenter()
            if col in (4, 5, 6) and val is not None:
                cell.number_format = "#,##0.00"
            if col == 6 and val is not None:
                cell.font = cfont(bold=True, color=GREEN_FG if val > 0 else RED_FG)
        ws.row_dimensions[r].height = 16

    # Totals
    tr = len(sorted_tx) + 3
    ws.cell(row=tr, column=3, value="TOTALS").font = hfont()
    ws.cell(row=tr, column=3).fill = solid(NAVY)
    ws.cell(row=tr, column=3).alignment = Alignment(horizontal="right", vertical="center")
    for col in (4, 5, 6):
        cl = get_column_letter(col)
        c  = ws.cell(row=tr, column=col, value=f"=SUM({cl}3:{cl}{tr-1})")
        c.font = hfont(); c.fill = solid(NAVY)
        c.number_format = "#,##0.00"; c.border = border()
        c.alignment = vcenter()
    ws.row_dimensions[tr].height = 20


def write_summary_sheet(ws, transactions):
    ws.title = "Summary by Category"
    cats = {}
    for tx in transactions:
        cat = tx.get("Category", "Other")
        amt = tx.get("Amount", 0)
        if cat not in cats:
            cats[cat] = {"in": 0.0, "out": 0.0, "count": 0}
        if amt > 0:
            cats[cat]["in"] += amt
        else:
            cats[cat]["out"] += abs(amt)
        cats[cat]["count"] += 1

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Spending Summary by Category"
    c.font  = hfont(size=13); c.fill = solid(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Category", "# Transactions", "Total In ($)", "Total Out ($)", "Net ($)"]
    widths  = [28, 18, 18, 18, 18]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = hfont(size=10); cell.fill = solid(BLUE)
        cell.alignment = center(); cell.border = border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for r, (cat, data) in enumerate(
        sorted(cats.items(), key=lambda x: x[1]["out"], reverse=True), start=3
    ):
        net = data["in"] - data["out"]
        bg  = solid(LIGHT if r % 2 == 0 else WHITE)
        vals = [cat, data["count"], data["in"], data["out"], net]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = cfont(); cell.fill = bg
            cell.border = border(); cell.alignment = vcenter()
            if col in (3, 4, 5):
                cell.number_format = "#,##0.00"
            if col == 5:
                cell.font = cfont(bold=True, color=GREEN_FG if net > 0 else RED_FG)
        ws.row_dimensions[r].height = 16

    tr = len(cats) + 3
    ws.cell(row=tr, column=1, value="GRAND TOTAL").font = hfont()
    ws.cell(row=tr, column=1).fill = solid(NAVY)
    ws.cell(row=tr, column=1).alignment = vcenter()
    for col in (2, 3, 4, 5):
        cl = get_column_letter(col)
        c  = ws.cell(row=tr, column=col, value=f"=SUM({cl}3:{cl}{tr-1})")
        c.font = hfont(); c.fill = solid(NAVY)
        c.number_format = "#,##0.00"; c.border = border()
        c.alignment = vcenter()
    ws.row_dimensions[tr].height = 20


def write_period_sheet(ws, transactions):
    ws.title = "By Statement Period"
    periods = {}
    for tx in transactions:
        p   = tx.get("Period", "Unknown")
        amt = tx.get("Amount", 0)
        if p not in periods:
            periods[p] = {"in": 0.0, "out": 0.0, "count": 0}
        if amt > 0:
            periods[p]["in"] += amt
        else:
            periods[p]["out"] += abs(amt)
        periods[p]["count"] += 1

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Activity by Statement Period"
    c.font  = hfont(size=13); c.fill = solid(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Statement Period", "# Transactions", "Total In ($)", "Total Out ($)", "Net ($)"]
    widths  = [42, 18, 18, 18, 18]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = hfont(size=10); cell.fill = solid(BLUE)
        cell.alignment = center(); cell.border = border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for r, (period, data) in enumerate(sorted(periods.items()), start=3):
        net = data["in"] - data["out"]
        bg  = solid(LIGHT if r % 2 == 0 else WHITE)
        vals = [period, data["count"], data["in"], data["out"], net]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = cfont(); cell.fill = bg
            cell.border = border(); cell.alignment = vcenter()
            if col in (3, 4, 5):
                cell.number_format = "#,##0.00"
            if col == 5:
                cell.font = cfont(bold=True, color=GREEN_FG if net > 0 else RED_FG)
        ws.row_dimensions[r].height = 16


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  RBC Statement Categorizer  v2")
    print("=" * 62)
    print()
    print("Drag your unlocked PDF folder into this window and press Enter:")
    print()
    folder_input = input("  Folder: ").strip().strip('"').strip("'")

    if not folder_input:
        print("\n  No folder provided.")
        input("Press Enter to close..."); return

    folder = Path(folder_input)
    if not folder.exists() or not folder.is_dir():
        print(f"\n  Folder not found: {folder}")
        input("Press Enter to close..."); return

    pdf_files = sorted(folder.glob("*.pdf"))
    if not pdf_files:
        print("\n  No PDF files found.")
        input("Press Enter to close..."); return

    print(f"\n  Found {len(pdf_files)} PDF file(s)\n")

    all_transactions = []
    errors = []

    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"  [{i}/{len(pdf_files)}] {pdf_path.name}")
        try:
            period_str, txs = extract_transactions_from_pdf(pdf_path)
            all_transactions.extend(txs)
            print(f"         → {len(txs)} transactions  ({period_str})")
        except Exception as e:
            errors.append((pdf_path.name, str(e)))
            print(f"         ✗ Error: {e}")

    if not all_transactions:
        print("\n  No transactions extracted.")
        print("  Tip: make sure you're using the _pwd_rmvd files (pre-redaction).")
        input("\nPress Enter to close..."); return

    print(f"\n  Total: {len(all_transactions)} transactions across {len(pdf_files)} statements")
    print("  Building Excel file...")

    wb = Workbook()
    write_transactions_sheet(wb.active, all_transactions)
    write_summary_sheet(wb.create_sheet(), all_transactions)
    write_period_sheet(wb.create_sheet(), all_transactions)

    output_path = folder / "RBC_Statement_Summary.xlsx"
    wb.save(str(output_path))

    print()
    print("=" * 62)
    print(f"  Saved to: {output_path}")
    print()
    print("  Sheets:")
    print("   • All Transactions   — every line item, sorted by date")
    print("   • Summary by Category — totals per spending category")
    print("   • By Statement Period — totals per monthly statement")
    if errors:
        print(f"\n  Files with errors ({len(errors)}):")
        for name, err in errors:
            print(f"   ✗ {name}: {err}")
    print("=" * 62)
    input("\nPress Enter to close...")


if __name__ == "__main__":
    main()
